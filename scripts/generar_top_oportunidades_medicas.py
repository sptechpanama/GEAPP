from __future__ import annotations

"""Genera un Top 50 estratégico de oportunidades médicas desde la capa analítica.

El informe usa la misma base analítica que consume Streamlit, recalcula la
unicidad de ficha por acto con el perfil moderado (score >= 90), cruza los
metadatos oficiales de MINSA/CTNI y produce exactamente cuatro hojas Excel.
"""

import argparse
from collections import defaultdict
import json
import math
import re
import sqlite3
import unicodedata
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


DETECTION_THRESHOLD = 90.0
TOP_N = 50
TOP_HIGHLIGHT = 5
MIN_CONTEXT_ACTS_FOR_GENERIC_NAME = 3
EXCLUDED_FICHAS = {
    "107135",
    "101792",
    "104747",
    "107260",
    "22998",
    "108528",
    "22287",
    "103152",
    "107861",
    # Ya revisada por el usuario; se excluye de todos los rankings futuros.
    "103169",
}
SHEET_NAMES = (
    "1_Historicas",
    "2_Nuevas_Potencial",
    "3_Barrera_Cero",
    "4_Actos_Desiertos",
)

MEDICAL_MARKERS = (
    "medico quirurgico",
    "dispositivo medico",
    "laboratorio",
    "odontolog",
    "imagenolog",
    "radiolog",
    "instrumental medico",
    "especialidades medicas",
)
PHARMA_MARKERS = (
    "medicamento",
    "medicamentosos",
    "farmaceut",
    "nutricion",
    "nutricional",
)

MASS_MARKERS = (
    "guante",
    "canula intravenosa",
    "aguja hipodermica",
    "jeringa",
    "gasa",
    "algodon",
    "mascarilla",
    "panal desechable",
    "papel termico",
)

PEROXIDE_COMPLETE_MARKERS = (
    "ciclo completo",
    "peroxido",
)

# Palabras que no aportan especificidad al nombre oficial de una ficha. Cuando
# el nombre queda reducido a un solo término útil (p. ej. CILINDRO, CUÑAS,
# ELECTRODO), una coincidencia literal no basta para atribuir el acto.
LOW_INFORMATION_STOPWORDS = {
    "de",
    "del",
    "la",
    "las",
    "el",
    "los",
    "para",
    "por",
    "con",
    "sin",
    "en",
    "al",
    "y",
    "o",
    "tipo",
}

CONTEXT_STOPWORDS = LOW_INFORMATION_STOPWORDS | {
    "una",
    "uno",
    "unos",
    "unas",
    "que",
    "como",
    "uso",
    "general",
    "medico",
    "medica",
    "medicos",
    "quirurgico",
    "quirurgica",
    "material",
    "materiales",
    "insumo",
    "insumos",
    "equipo",
    "equipos",
    "dispositivo",
    "dispositivos",
    "instrumental",
    "especialidad",
}

# Raíces que vuelven suficientemente específica una frase corta de dos
# términos (por ejemplo, BOMBA DE INFUSIÓN). Una frase corta sin estas señales
# se valida como nombre genérico para evitar colisiones fuera del ámbito médico.
DISTINCTIVE_MEDICAL_PREFIXES = (
    "anest",
    "biolog",
    "biops",
    "bronco",
    "cardiac",
    "catet",
    "dialisis",
    "endosc",
    "esteril",
    "hemat",
    "hemodial",
    "hemost",
    "infus",
    "lapar",
    "odontolog",
    "oxigen",
    "oxim",
    "protes",
    "quirurg",
    "radiolog",
    "sutura",
    "terapeut",
    "toracic",
    "traque",
    "ventil",
)


@dataclass(frozen=True)
class InputPaths:
    analytics_db: Path
    ctni_db: Path
    operational_db: Path


def clean_text(value: object) -> str:
    text = str(value if value is not None else "").strip()
    if text.lower() in {"", "nan", "none", "null", "<na>"}:
        return ""
    return re.sub(r"\s+", " ", text)


def normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value).lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", text)).strip()


def _meaningful_name_tokens(value: object) -> tuple[str, ...]:
    return tuple(
        token
        for token in normalize_text(value).split()
        if len(token) >= 3 and token not in LOW_INFORMATION_STOPWORDS
    )


def _is_low_information_name(value: object) -> bool:
    """Identifica nombres cuyo literal aislado tiene alto riesgo de colisión."""
    tokens = _meaningful_name_tokens(value)
    if len(tokens) <= 1:
        return True
    if len(tokens) == 2:
        return not any(
            token.startswith(prefix)
            for token in tokens
            for prefix in DISTINCTIVE_MEDICAL_PREFIXES
        )
    return False


def _validation_context_tokens(row: pd.Series) -> tuple[str, ...]:
    """Extrae términos técnicos secundarios para validar un alias genérico."""
    official_name = clean_text(row.get("nombre_ficha")) or clean_text(
        row.get("ctni_producto_oficial")
    )
    name_tokens = set(_meaningful_name_tokens(official_name))
    raw_context = " ".join(
        clean_text(row.get(column))
        for column in (
            "descripcion",
            "tipo_producto",
            "especialidad",
            "area",
            "ctni_subcomite",
        )
    )
    tokens: list[str] = []
    for token in normalize_text(raw_context).split():
        if (
            len(token) < 4
            or token in CONTEXT_STOPWORDS
            or token in name_tokens
            or token in tokens
        ):
            continue
        tokens.append(token)
    return tuple(tokens[:24])


def _context_hit_count(search_text: object, context_tokens: object) -> int:
    words = set(normalize_text(search_text).split())
    if not words or not isinstance(context_tokens, (tuple, list)):
        return 0
    hits = 0
    for token in context_tokens:
        normalized = normalize_text(token)
        if not normalized:
            continue
        prefix = normalized[:6] if len(normalized) >= 6 else normalized
        if any(word == normalized or word.startswith(prefix) for word in words):
            hits += 1
    return hits


def _contains_explicit_ficha(row: pd.Series) -> bool:
    method = normalize_text(row.get("detection_method"))
    if "codigo" in method:
        return True
    ficha = normalize_ficha(row.get("ficha"))
    if not ficha:
        return False
    evidence = normalize_text(
        f"{clean_text(row.get('detection_evidence'))} {clean_text(row.get('search_text_norm'))}"
    )
    return bool(re.search(rf"(?<!\d){re.escape(ficha)}(?!\d)", evidence))


def normalize_ficha(value: object) -> str:
    match = re.search(r"\d+", clean_text(value))
    return (match.group(0).lstrip("0") or "0") if match else ""


def normalize_requirement(value: object) -> str:
    """Devuelve ``si``, ``no`` o vacío sin inferir requisitos desconocidos."""
    text = normalize_text(value)
    if text in {"si", "s", "true", "1"}:
        return "si"
    if text in {"no", "n", "false", "0", "no aplica", "no aplica no"}:
        return "no"
    return ""


def parse_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value) if math.isfinite(float(value)) else 0.0
    text = clean_text(value).replace("$", "").replace(" ", "")
    if not text:
        return 0.0
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        tail = text.rsplit(",", 1)[-1]
        text = text.replace(",", ".") if len(tail) <= 2 else text.replace(",", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    try:
        parsed = float(text)
    except (TypeError, ValueError):
        return 0.0
    return parsed if math.isfinite(parsed) else 0.0


def _quantile(values: Iterable[float], q: float) -> float:
    series = pd.Series([float(value) for value in values if float(value) > 0], dtype="float64")
    return float(series.quantile(q)) if not series.empty else 0.0


def parse_iso_date(value: object) -> pd.Timestamp:
    return pd.to_datetime(clean_text(value), errors="coerce")


def _read_ctni_catalog(ctni_db: Path) -> pd.DataFrame:
    """Consolida el histórico CTNI por ficha y conserva el nombre oficial completo."""
    rows: list[dict[str, object]] = []
    with sqlite3.connect(ctni_db) as connection:
        payloads = connection.execute(
            "SELECT payload_json FROM records WHERE categoria = 'fichas'"
        ).fetchall()
    for (payload_raw,) in payloads:
        try:
            payload = json.loads(payload_raw)
        except (TypeError, ValueError, json.JSONDecodeError):
            continue
        ficha = normalize_ficha(payload.get("numero_ficha"))
        event_date = parse_iso_date(payload.get("fecha"))
        if not ficha or pd.isna(event_date):
            continue
        rows.append(
            {
                "ficha": ficha,
                "ctni_fecha": event_date,
                "ctni_accion": clean_text(payload.get("accion")),
                "ctni_producto": clean_text(payload.get("producto")),
                "ctni_subcomite": clean_text(payload.get("subcomite")),
                "ctni_enlace": clean_text(payload.get("enlace_oficial")),
            }
        )
    if not rows:
        return pd.DataFrame(
            columns=[
                "ficha",
                "ctni_primera_fecha",
                "ctni_ultima_fecha",
                "ctni_ultima_accion",
                "ctni_producto_oficial",
                "ctni_subcomite",
                "ctni_enlace",
            ]
        )
    events = pd.DataFrame(rows).sort_values(["ficha", "ctni_fecha"])
    first_dates = events.groupby("ficha", as_index=False)["ctni_fecha"].min().rename(
        columns={"ctni_fecha": "ctni_primera_fecha"}
    )
    latest = events.groupby("ficha", as_index=False).tail(1).rename(
        columns={
            "ctni_fecha": "ctni_ultima_fecha",
            "ctni_accion": "ctni_ultima_accion",
            "ctni_producto": "ctni_producto_oficial",
        }
    )
    latest = latest[
        [
            "ficha",
            "ctni_ultima_fecha",
            "ctni_ultima_accion",
            "ctni_producto_oficial",
            "ctni_subcomite",
            "ctni_enlace",
        ]
    ]
    return first_dates.merge(latest, on="ficha", how="left")


def _load_source(paths: InputPaths) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, str]]:
    with sqlite3.connect(paths.analytics_db) as connection:
        facts = pd.read_sql_query(
            """
            SELECT *
            FROM intel_actos_fichas
            WHERE detection_score >= ?
            """,
            connection,
            params=(DETECTION_THRESHOLD,),
        )
        metadata = pd.read_sql_query("SELECT * FROM intel_ficha_metadata", connection)
        build_metadata = dict(
            connection.execute("SELECT key, value FROM intel_build_metadata").fetchall()
        )
    ctni = _read_ctni_catalog(paths.ctni_db)
    for frame in (facts, metadata, ctni):
        if "ficha" in frame.columns:
            frame["ficha"] = frame["ficha"].map(normalize_ficha)
    metadata = metadata.drop_duplicates("ficha", keep="last")
    metadata = metadata.merge(ctni, on="ficha", how="outer")
    return facts, metadata, build_metadata


def _load_price_intelligence(paths: InputPaths, known_fichas: set[str]) -> pd.DataFrame:
    """Resume presión de ofertas y precios unitarios observados por ficha.

    Las ofertas se comparan únicamente en actos de ficha única para evitar
    atribuir a una ficha el total de un acto mixto. Los precios unitarios se
    extraen de ``items_json`` y se agrupan por la unidad de medida dominante;
    no se mezclan cajas con unidades.
    """
    with sqlite3.connect(paths.analytics_db) as connection:
        bids = pd.read_sql_query(
            """
            SELECT
                f.ficha,
                f.acto_key,
                p.offered_amount,
                p.is_winner,
                f.reference_amount_context,
                f.participant_count
            FROM intel_actos_fichas AS f
            JOIN intel_acto_proponentes AS p ON p.acto_key = f.acto_key
            WHERE f.detection_score >= ?
              AND f.is_unique_ficha = 1
              AND p.offered_amount > 0
              AND f.reference_amount_context > 0
            """,
            connection,
            params=(DETECTION_THRESHOLD,),
        )
    bid_rows: list[dict[str, object]] = []
    if not bids.empty:
        bids["ficha"] = bids["ficha"].map(normalize_ficha)
        bids["ratio"] = bids["offered_amount"] / bids["reference_amount_context"]
        bids = bids[
            bids["ficha"].isin(known_fichas)
            & bids["ratio"].between(0.03, 3.0, inclusive="both")
        ].copy()
        for ficha, group in bids.groupby("ficha", sort=False):
            winners = group[group["is_winner"].fillna(0).astype(int).eq(1)]
            ratio_source = winners["ratio"] if not winners.empty else group["ratio"]
            bid_rows.append(
                {
                    "ficha": ficha,
                    "propuestas_observadas": int(len(group)),
                    "actos_precio_observados": int(group["acto_key"].nunique()),
                    "oferta_total_mediana": float(group["offered_amount"].median()),
                    "oferta_total_p25": float(group["offered_amount"].quantile(0.25)),
                    "oferta_total_p75": float(group["offered_amount"].quantile(0.75)),
                    "ratio_oferta_referencia_mediana": float(ratio_source.median()),
                    "ratio_oferta_referencia_p25": float(ratio_source.quantile(0.25)),
                }
            )

    unit_values: dict[tuple[str, str], list[float]] = defaultdict(list)
    with sqlite3.connect(paths.operational_db) as connection:
        cursor = connection.execute(
            """
            SELECT ficha_detectada, fichas_detectadas_json, items_json
            FROM actos_publicos
            WHERE items_json IS NOT NULL AND TRIM(items_json) NOT IN ('', '[]', 'null')
            """
        )
        for ficha_detectada, fichas_json, items_raw in cursor:
            act_fichas = {
                normalize_ficha(token)
                for token in re.findall(r"(?<!\d)\d{3,8}(?!\d)", clean_text(ficha_detectada))
            }
            if fichas_json:
                try:
                    parsed_fichas = json.loads(fichas_json)
                except (TypeError, ValueError, json.JSONDecodeError):
                    parsed_fichas = []
                if isinstance(parsed_fichas, dict):
                    parsed_fichas = list(parsed_fichas)
                if isinstance(parsed_fichas, list):
                    for value in parsed_fichas:
                        raw_value = value.get("ficha") if isinstance(value, dict) else value
                        normalized = normalize_ficha(raw_value)
                        if normalized:
                            act_fichas.add(normalized)
            act_fichas &= known_fichas
            if not act_fichas:
                continue
            try:
                items = json.loads(items_raw)
            except (TypeError, ValueError, json.JSONDecodeError):
                continue
            if not isinstance(items, list):
                continue
            for item in items:
                if not isinstance(item, dict):
                    continue
                description = clean_text(item.get("descripcion"))
                explicit = {
                    ficha
                    for ficha in act_fichas
                    if re.search(rf"(?<!\d){re.escape(ficha)}(?!\d)", description)
                }
                if len(explicit) == 1:
                    assigned = next(iter(explicit))
                elif len(explicit) == 0 and len(act_fichas) == 1:
                    assigned = next(iter(act_fichas))
                else:
                    continue
                unit_price = parse_number(item.get("precio_referencia_unitario"))
                if unit_price <= 0 or unit_price > 10_000_000:
                    continue
                unit = normalize_text(item.get("unidad")) or "sin unidad"
                unit_values[(assigned, unit)].append(unit_price)

    by_ficha_units: dict[str, list[tuple[str, list[float]]]] = defaultdict(list)
    for (ficha, unit), values in unit_values.items():
        by_ficha_units[ficha].append((unit, values))
    unit_rows: list[dict[str, object]] = []
    for ficha, groups in by_ficha_units.items():
        unit, values = max(groups, key=lambda pair: (len(pair[1]), pair[0]))
        unit_rows.append(
            {
                "ficha": ficha,
                "unidad_precio_dominante": unit,
                "precios_unitarios_observados": len(values),
                "precio_unitario_ref_p25": _quantile(values, 0.25),
                "precio_unitario_ref_mediana": _quantile(values, 0.50),
                "precio_unitario_ref_p75": _quantile(values, 0.75),
            }
        )

    bid_frame = pd.DataFrame(bid_rows)
    unit_frame = pd.DataFrame(unit_rows)
    if bid_frame.empty and unit_frame.empty:
        return pd.DataFrame(columns=["ficha"])
    if bid_frame.empty:
        return unit_frame
    if unit_frame.empty:
        return bid_frame
    return bid_frame.merge(unit_frame, on="ficha", how="outer")


def _derive_medical_scope(metadata: pd.DataFrame) -> pd.DataFrame:
    result = metadata.copy()
    for column in (
        "nombre_ficha",
        "descripcion",
        "area",
        "tipo_producto",
        "especialidad",
        "ctni_producto_oficial",
        "ctni_subcomite",
    ):
        if column not in result.columns:
            result[column] = ""
        result[column] = result[column].map(clean_text)

    result["descripcion_oficial"] = result.apply(
        lambda row: clean_text(row.get("ctni_producto_oficial"))
        or clean_text(row.get("nombre_ficha"))
        or clean_text(row.get("descripcion"))
        or f"Ficha {row.get('ficha', '')}",
        axis=1,
    )
    result["nombre_validacion"] = result.apply(
        lambda row: clean_text(row.get("nombre_ficha"))
        or clean_text(row.get("ctni_producto_oficial"))
        or clean_text(row.get("descripcion_oficial")),
        axis=1,
    )
    result["alias_baja_informacion"] = result["nombre_validacion"].map(
        _is_low_information_name
    )
    result["tokens_contexto_validacion"] = result.apply(
        _validation_context_tokens,
        axis=1,
    )
    result["clasificacion_oficial"] = result.apply(
        lambda row: clean_text(row.get("ctni_subcomite"))
        or clean_text(row.get("area"))
        or clean_text(row.get("tipo_producto")),
        axis=1,
    )
    taxonomy = result.apply(
        lambda row: normalize_text(
            " ".join(
                clean_text(row.get(column))
                for column in (
                    "ctni_subcomite",
                    "area",
                    "tipo_producto",
                    "especialidad",
                    "ctni_producto_oficial",
                    "nombre_ficha",
                )
            )
        ),
        axis=1,
    )
    is_pharma = taxonomy.map(lambda text: any(marker in text for marker in PHARMA_MARKERS))
    is_medical = taxonomy.map(lambda text: any(marker in text for marker in MEDICAL_MARKERS))
    result["es_universo_medico"] = is_medical & ~is_pharma
    result["motivo_clasificacion"] = ""
    result.loc[is_medical & ~is_pharma, "motivo_clasificacion"] = "Taxonomía oficial médica"
    result.loc[is_pharma, "motivo_clasificacion"] = "Excluida: medicamento/nutrición"
    result.loc[~is_medical & ~is_pharma, "motivo_clasificacion"] = "Excluida: clasificación no confirmada"
    return result


def _prepare_facts(facts: pd.DataFrame, metadata: pd.DataFrame) -> pd.DataFrame:
    result = facts.copy()
    result["ficha"] = result["ficha"].map(normalize_ficha)
    result["profile_ficha_count"] = result.groupby("acto_key")["ficha"].transform("nunique")
    result["es_ficha_unica"] = result["profile_ficha_count"].eq(1)
    for column in (
        "reference_amount_context",
        "award_amount_context",
        "participant_count",
        "detection_score",
    ):
        result[column] = pd.to_numeric(result[column], errors="coerce").fillna(0.0)
    for column in ("publication_date", "celebration_date", "award_date", "update_date"):
        result[column] = pd.to_datetime(result[column], errors="coerce")
    result["fecha_analisis"] = (
        result["publication_date"]
        .fillna(result["celebration_date"])
        .fillna(result["award_date"])
        .fillna(result["update_date"])
    )
    result["estado_norm"] = result["estado"].map(normalize_text)
    result["es_desierto"] = result["estado_norm"].str.contains("desiert", na=False)
    result["es_adjudicado"] = result["estado_norm"].str.contains("adjudic", na=False)
    result = result.merge(
        metadata[
            [
                "ficha",
                "descripcion_oficial",
                "clasificacion_oficial",
                "es_universo_medico",
                "tiene_ct",
                "registro_sanitario",
                "clase_riesgo",
                "enlace_minsa",
                "ctni_enlace",
                "ctni_primera_fecha",
                "ctni_ultima_fecha",
                "ctni_ultima_accion",
                "nombre_validacion",
                "alias_baja_informacion",
                "tokens_contexto_validacion",
            ]
        ],
        on="ficha",
        how="left",
        validate="many_to_one",
    )
    result["rs_norm"] = result["registro_sanitario"].map(normalize_requirement)
    result["ct_norm"] = result["tiene_ct"].map(normalize_requirement)
    product_norm = result["descripcion_oficial"].map(normalize_text)
    peroxide_complete = product_norm.map(
        lambda text: all(marker in text for marker in PEROXIDE_COMPLETE_MARKERS)
    )
    scoped = result[
        result["es_universo_medico"].fillna(False)
        & ~result["ficha"].isin(EXCLUDED_FICHAS)
        & result["rs_norm"].eq("no")
        & ~peroxide_complete
    ].copy()
    scoped["alias_baja_informacion"] = scoped["alias_baja_informacion"].fillna(False).astype(bool)
    scoped["evidencia_codigo_explicito"] = scoped.apply(_contains_explicit_ficha, axis=1)
    scoped["coincidencias_contexto"] = scoped.apply(
        lambda row: _context_hit_count(
            row.get("search_text_norm"), row.get("tokens_contexto_validacion")
        ),
        axis=1,
    )
    scoped["evidencia_contexto_reforzado"] = scoped["coincidencias_contexto"].ge(2)
    scoped["relacion_validada"] = (
        ~scoped["alias_baja_informacion"]
        | scoped["evidencia_codigo_explicito"]
        | scoped["evidencia_contexto_reforzado"]
    )

    raw_counts = scoped.groupby("ficha")["acto_key"].nunique()
    validated = scoped[scoped["relacion_validada"]].copy()
    validated_counts = validated.groupby("ficha")["acto_key"].nunique()
    explicit_counts = validated.loc[
        validated["evidencia_codigo_explicito"]
    ].groupby("ficha")["acto_key"].nunique()
    contextual_counts = validated.loc[
        ~validated["evidencia_codigo_explicito"]
        & validated["evidencia_contexto_reforzado"]
    ].groupby("ficha")["acto_key"].nunique()
    validated["actos_detectados_brutos"] = validated["ficha"].map(raw_counts).fillna(0).astype(int)
    validated["actos_validados"] = validated["ficha"].map(validated_counts).fillna(0).astype(int)
    validated["actos_codigo_explicito"] = validated["ficha"].map(explicit_counts).fillna(0).astype(int)
    validated["actos_contexto_reforzado"] = validated["ficha"].map(contextual_counts).fillna(0).astype(int)
    validated["actos_descartados_ambiguos"] = (
        validated["actos_detectados_brutos"] - validated["actos_validados"]
    ).clip(lower=0)
    validated.attrs["quality_dropped_relations"] = int((~scoped["relacion_validada"]).sum())
    validated.attrs["quality_dropped_acts"] = int(
        scoped.loc[~scoped["relacion_validada"], "acto_key"].nunique()
    )
    return validated


def _percentile(series: pd.Series, *, higher_is_better: bool = True) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce").fillna(0.0)
    if values.empty:
        return values
    if values.nunique(dropna=False) <= 1:
        return pd.Series(50.0, index=values.index)
    ranked = values.rank(method="average", pct=True) * 100.0
    return ranked if higher_is_better else 100.0 - ranked + (100.0 / len(values))


def _competition_label(median_participants: float, avg_participants: float) -> str:
    reference = median_participants if median_participants > 0 else avg_participants
    if reference <= 0:
        return "Sin datos de proponentes"
    if reference <= 1.25:
        level = "Baja"
    elif reference <= 2.5:
        level = "Media"
    else:
        level = "Alta"
    return f"{level} (mediana {median_participants:.1f}; promedio {avg_participants:.1f})"


def _requirements(ct_value: object, rs_value: object, risk_class: object) -> str:
    ct = clean_text(ct_value) or "Sin dato"
    rs = clean_text(rs_value) or "Sin dato"
    risk = clean_text(risk_class) or "Sin clase"
    return f"CT: {ct} | RS: {rs} | Clase: {risk}"


def _aggregate(facts: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for ficha, group in facts.groupby("ficha", sort=False):
        positive_participants = group.loc[group["participant_count"] > 0, "participant_count"]
        acts = int(group["acto_key"].nunique())
        unique_group = group[group["es_ficha_unica"]]
        adjudicated_acts = int(group.loc[group["es_adjudicado"], "acto_key"].nunique())
        deserted_acts = int(group.loc[group["es_desierto"], "acto_key"].nunique())
        first = group.iloc[0]
        total_amount = float(group["reference_amount_context"].clip(lower=0).sum())
        unique_amount = float(unique_group["reference_amount_context"].clip(lower=0).sum())
        award_amount = float(group["award_amount_context"].clip(lower=0).sum())
        avg_participants = float(positive_participants.mean()) if not positive_participants.empty else 0.0
        median_participants = float(positive_participants.median()) if not positive_participants.empty else 0.0
        rows.append(
            {
                "ficha": ficha,
                "descripcion_oficial": clean_text(first.get("descripcion_oficial")),
                "clasificacion_oficial": clean_text(first.get("clasificacion_oficial")),
                "actos": acts,
                "actos_ficha_unica": int(unique_group["acto_key"].nunique()),
                "monto_total": total_amount,
                "monto_ficha_unica": unique_amount,
                "monto_adjudicado": award_amount,
                "promedio_por_acto": total_amount / acts if acts else 0.0,
                "meses_activos": int(group["fecha_analisis"].dt.to_period("M").nunique()),
                "primera_fecha": group["fecha_analisis"].min(),
                "ultima_fecha": group["fecha_analisis"].max(),
                "actos_adjudicados": adjudicated_acts,
                "actos_desiertos": deserted_acts,
                "pct_adjudicado": adjudicated_acts / acts if acts else 0.0,
                "pct_desierto": deserted_acts / acts if acts else 0.0,
                "participantes_promedio": avg_participants,
                "participantes_mediana": median_participants,
                "competencia": _competition_label(median_participants, avg_participants),
                "estatus_adjudicacion": (
                    f"Adjudicados {adjudicated_acts}/{acts} ({adjudicated_acts / acts:.1%}); "
                    f"desiertos {deserted_acts}/{acts} ({deserted_acts / acts:.1%})"
                    if acts
                    else "Sin actos"
                ),
                "tiene_ct": clean_text(first.get("tiene_ct")),
                "registro_sanitario": clean_text(first.get("registro_sanitario")),
                "clase_riesgo": clean_text(first.get("clase_riesgo")),
                "requisitos": _requirements(
                    first.get("tiene_ct"), first.get("registro_sanitario"), first.get("clase_riesgo")
                ),
                "enlace_minsa": clean_text(first.get("enlace_minsa"))
                or clean_text(first.get("ctni_enlace")),
                "ctni_primera_fecha": first.get("ctni_primera_fecha"),
                "ctni_ultima_fecha": first.get("ctni_ultima_fecha"),
                "ctni_ultima_accion": clean_text(first.get("ctni_ultima_accion")),
                "confianza_deteccion": float(group["detection_score"].mean()),
                "alias_baja_informacion": bool(first.get("alias_baja_informacion")),
                "actos_detectados_brutos": int(first.get("actos_detectados_brutos") or acts),
                "actos_descartados_ambiguos": int(first.get("actos_descartados_ambiguos") or 0),
                "actos_codigo_explicito": int(first.get("actos_codigo_explicito") or 0),
                "actos_contexto_reforzado": int(first.get("actos_contexto_reforzado") or 0),
            }
        )
    result = pd.DataFrame(rows)
    if result.empty:
        return result
    result["validacion_deteccion"] = result.apply(
        lambda row: (
            "Frase oficial específica"
            if not bool(row.get("alias_baja_informacion"))
            else (
                f"Nombre genérico validado con código oficial en "
                f"{int(row.get('actos_codigo_explicito') or 0)} acto(s)"
                if int(row.get("actos_codigo_explicito") or 0) > 0
                else (
                    f"Nombre genérico validado por contexto técnico en "
                    f"{int(row.get('actos_contexto_reforzado') or 0)} acto(s)"
                )
            )
        ),
        axis=1,
    )
    return result


def _eligible_detection_quality(frame: pd.DataFrame) -> pd.DataFrame:
    """Excluye alias genéricos sin evidencia suficiente para un ranking comercial."""
    if frame.empty:
        return frame.copy()
    generic_supported = (
        frame["actos_codigo_explicito"].gt(0)
        | frame["actos_contexto_reforzado"].ge(MIN_CONTEXT_ACTS_FOR_GENERIC_NAME)
    )
    return frame[
        ~frame["alias_baja_informacion"].fillna(False) | generic_supported
    ].copy()


def _enrich_price_viability(frame: pd.DataFrame, price_intelligence: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    if not price_intelligence.empty:
        result = result.merge(price_intelligence, on="ficha", how="left", validate="one_to_one")
    numeric_columns = (
        "propuestas_observadas",
        "actos_precio_observados",
        "oferta_total_mediana",
        "oferta_total_p25",
        "oferta_total_p75",
        "ratio_oferta_referencia_mediana",
        "ratio_oferta_referencia_p25",
        "precios_unitarios_observados",
        "precio_unitario_ref_p25",
        "precio_unitario_ref_mediana",
        "precio_unitario_ref_p75",
    )
    for column in numeric_columns:
        if column not in result.columns:
            result[column] = 0.0
        result[column] = pd.to_numeric(result[column], errors="coerce").fillna(0.0)
    if "unidad_precio_dominante" not in result.columns:
        result["unidad_precio_dominante"] = ""
    result["unidad_precio_dominante"] = result["unidad_precio_dominante"].map(clean_text)
    description_norm = result["descripcion_oficial"].map(normalize_text)
    result["producto_masivo"] = description_norm.map(
        lambda text: any(marker in text for marker in MASS_MARKERS)
    )

    def margin_score(row: pd.Series) -> float:
        ratio = float(row.get("ratio_oferta_referencia_mediana") or 0.0)
        if ratio > 0:
            ratio_score = min(max((ratio - 0.45) / 0.55 * 100.0, 0.0), 100.0)
        else:
            ratio_score = 50.0
        participants = float(row.get("participantes_mediana") or 0.0)
        competition_score = 75.0 if participants <= 0 else max(10.0, 100.0 - (participants - 1.0) * 25.0)
        score = ratio_score * 0.75 + competition_score * 0.25
        unit_price = float(row.get("precio_unitario_ref_mediana") or 0.0)
        if bool(row.get("producto_masivo")):
            if 0 < unit_price <= 0.10:
                score -= 30.0
            elif unit_price <= 0.50:
                score -= 18.0
            elif unit_price <= 2.00:
                score -= 8.0
        return min(max(score, 0.0), 100.0)

    result["score_margen"] = result.apply(margin_score, axis=1)

    def pressure(row: pd.Series) -> str:
        ratio = float(row.get("ratio_oferta_referencia_mediana") or 0.0)
        unit = float(row.get("precio_unitario_ref_mediana") or 0.0)
        mass = bool(row.get("producto_masivo"))
        if ratio <= 0:
            return "Sin muestra suficiente"
        if ratio < 0.60 or (mass and 0 < unit <= 0.10):
            return "Muy alta"
        if ratio < 0.78 or (mass and 0 < unit <= 0.50):
            return "Alta"
        if ratio < 0.93:
            return "Media"
        return "Baja"

    result["presion_precio"] = result.apply(pressure, axis=1)

    def viability(row: pd.Series) -> str:
        observations = int(row.get("propuestas_observadas") or 0)
        score = float(row.get("score_margen") or 0.0)
        if observations <= 0:
            return "Por verificar: sin muestra comparable"
        if score < 35:
            return "Baja: no priorizar sin fabricante directo"
        if score < 55:
            return "Condicionada: validar costo puesto"
        if score < 75:
            return "Media: margen posible con compra directa"
        return "Favorable preliminar"

    result["viabilidad_margen"] = result.apply(viability, axis=1)
    ratio_for_target = result["ratio_oferta_referencia_mediana"].where(
        result["ratio_oferta_referencia_mediana"].gt(0), 0.85
    ).clip(lower=0.30, upper=1.10)
    result["costo_objetivo_unitario"] = (
        result["precio_unitario_ref_mediana"] * ratio_for_target * 0.75
    )

    def commercial_note(row: pd.Series) -> str:
        ratio = float(row.get("ratio_oferta_referencia_mediana") or 0.0)
        proposals = int(row.get("propuestas_observadas") or 0)
        unit = float(row.get("precio_unitario_ref_mediana") or 0.0)
        unit_name = clean_text(row.get("unidad_precio_dominante")) or "unidad declarada"
        pieces: list[str] = []
        if proposals:
            pieces.append(
                f"{proposals} propuestas comparables; oferta ganadora/participación típica ≈ {ratio:.0%} de la referencia"
            )
        else:
            pieces.append("sin suficientes propuestas comparables de ficha única")
        if unit > 0:
            pieces.append(
                f"referencia mediana USD {unit:,.4f} por {unit_name}; costo puesto objetivo ≤ USD {float(row['costo_objetivo_unitario']):,.4f} para 25% bruto"
            )
        if bool(row.get("producto_masivo")):
            pieces.append("producto masivo: exigir cotización directa de fábrica y validar empaque")
        return "; ".join(pieces) + "."

    result["nota_viabilidad_precio"] = result.apply(commercial_note, axis=1)
    return result


def _score_historical(frame: pd.DataFrame) -> pd.DataFrame:
    result = _eligible_detection_quality(frame)
    result["score"] = (
        _percentile(result["actos"]) * 0.25
        + _percentile(result["actos_ficha_unica"]) * 0.20
        + _percentile(result["monto_ficha_unica"]) * 0.25
        + _percentile(result["monto_total"]) * 0.10
        + _percentile(result["meses_activos"]) * 0.10
        + result["score_margen"] * 0.10
    )
    return result.sort_values(["score", "monto_ficha_unica", "actos"], ascending=False)


def _score_new(
    facts: pd.DataFrame,
    cutoff: pd.Timestamp,
    price_intelligence: pd.DataFrame,
) -> pd.DataFrame:
    # Una ficha nueva no puede heredar actos anteriores a su creación. Las
    # coincidencias semánticas históricas son útiles para otros análisis, pero
    # aquí inflarían artificialmente la adopción real del código recién creado.
    eligible_facts = facts[
        facts["ctni_primera_fecha"].notna()
        & (facts["ctni_primera_fecha"] >= cutoff)
        & facts["fecha_analisis"].notna()
        & (facts["fecha_analisis"] >= facts["ctni_primera_fecha"])
    ].copy()
    result = _eligible_detection_quality(
        _enrich_price_viability(_aggregate(eligible_facts), price_intelligence)
    )
    result = result[
        # Para ampliar de Top 20 a Top 50 conservamos al menos una señal real
        # posterior a la creación y exigimos que sea un acto de ficha única.
        (result["actos"] >= 1)
        & (result["actos_ficha_unica"] >= 1)
        & (result["monto_ficha_unica"] > 0)
    ].copy()
    age_months = ((pd.Timestamp.today().normalize() - result["ctni_primera_fecha"]).dt.days / 30.44).clip(lower=1)
    result["dinamismo_mensual"] = result["actos"] / age_months
    result["score"] = (
        _percentile(result["actos"]) * 0.22
        + _percentile(result["monto_ficha_unica"]) * 0.23
        + _percentile(result["monto_total"]) * 0.10
        + _percentile(result["participantes_mediana"], higher_is_better=False) * 0.15
        + _percentile(result["dinamismo_mensual"]) * 0.15
        + result["score_margen"] * 0.15
    )
    return result.sort_values(["score", "monto_ficha_unica", "actos"], ascending=False)


def _score_barrier_zero(frame: pd.DataFrame) -> pd.DataFrame:
    eligible = _eligible_detection_quality(frame)
    result = eligible[
        eligible["tiene_ct"].map(normalize_requirement).eq("no")
        & eligible["registro_sanitario"].map(normalize_requirement).eq("no")
    ].copy()
    result["score"] = (
        _percentile(result["actos"]) * 0.25
        + _percentile(result["actos_ficha_unica"]) * 0.20
        + _percentile(result["monto_ficha_unica"]) * 0.25
        + _percentile(result["monto_total"]) * 0.08
        + _percentile(result["participantes_mediana"], higher_is_better=False) * 0.07
        + result["score_margen"] * 0.15
    )
    return result.sort_values(["score", "monto_ficha_unica", "actos"], ascending=False)


def _score_deserted(facts: pd.DataFrame, price_intelligence: pd.DataFrame) -> pd.DataFrame:
    deserted = facts[facts["es_desierto"]].copy()
    result = _eligible_detection_quality(
        _enrich_price_viability(_aggregate(deserted), price_intelligence)
    )
    result = result[result["actos"] >= 1].copy()
    result["score"] = (
        _percentile(result["actos"]) * 0.40
        + _percentile(result["monto_total"]) * 0.25
        + _percentile(result["actos_ficha_unica"]) * 0.10
        + _percentile(result["meses_activos"]) * 0.10
        + result["score_margen"] * 0.15
    )
    return result.sort_values(["score", "monto_total", "actos"], ascending=False)


def _reason(row: pd.Series, category: str, rank: int) -> str:
    prefix = "TOP 5: " if rank <= 5 else ""
    if category == "historical":
        return (
            f"{prefix}{int(row['actos'])} actos en {int(row['meses_activos'])} meses activos; "
            f"USD {row['monto_ficha_unica']:,.2f} en actos de ficha única y "
            f"USD {row['monto_total']:,.2f} de mercado relacionado."
        )
    if category == "new":
        event_date = pd.to_datetime(row.get("ctni_primera_fecha"), errors="coerce")
        event_label = event_date.strftime("%Y-%m-%d") if pd.notna(event_date) else "sin fecha"
        return (
            f"{prefix}Ficha reciente desde {event_label}; {int(row['actos'])} actos, "
            f"USD {row['monto_ficha_unica']:,.2f} de ficha única y competencia {row['competencia'].lower()}."
        )
    if category == "barrier":
        return (
            f"{prefix}Entrada regulatoria directa (CT No y RS No), {int(row['actos'])} actos y "
            f"USD {row['monto_ficha_unica']:,.2f} en demanda de ficha única."
        )
    return (
        f"{prefix}{int(row['actos'])} actos desiertos por USD {row['monto_total']:,.2f}; "
        f"señal de demanda no satisfecha recurrente en {int(row['meses_activos'])} meses."
    )


def _to_export(frame: pd.DataFrame, category: str, top_n: int = TOP_N) -> pd.DataFrame:
    top = frame.head(top_n).copy().reset_index(drop=True)
    top.insert(0, "Ranking", range(1, len(top) + 1))
    top["Prioridad"] = top["Ranking"].map(
        lambda value: f"⭐ TOP {TOP_HIGHLIGHT}" if value <= TOP_HIGHLIGHT else f"Top {TOP_N}"
    )
    top["Por qué destaca"] = [
        _reason(row, category, int(row["Ranking"])) for _, row in top.iterrows()
    ]
    top["Rango unitario de referencia"] = top.apply(
        lambda row: (
            f"USD {float(row.get('precio_unitario_ref_p25') or 0):,.4f} – "
            f"USD {float(row.get('precio_unitario_ref_p75') or 0):,.4f}"
            if float(row.get("precio_unitario_ref_mediana") or 0) > 0
            else "Sin muestra suficiente"
        ),
        axis=1,
    )
    if category == "new":
        top["Fecha publicación/creación"] = pd.to_datetime(
            top["ctni_primera_fecha"], errors="coerce"
        ).dt.date
        top["Acción CTNI más reciente"] = top["ctni_ultima_accion"]
    columns = [
        "Ranking",
        "Prioridad",
        "ficha",
        "descripcion_oficial",
        "validacion_deteccion",
        "actos_descartados_ambiguos",
    ]
    if category == "new":
        columns.extend(["Fecha publicación/creación", "Acción CTNI más reciente"])
    columns.extend(
        [
            "actos",
            "actos_ficha_unica",
            "monto_total",
            "monto_ficha_unica",
            "promedio_por_acto",
            "competencia",
            "participantes_promedio",
            "participantes_mediana",
            "estatus_adjudicacion",
            "requisitos",
            "clasificacion_oficial",
            "clase_riesgo",
            "presion_precio",
            "viabilidad_margen",
            "propuestas_observadas",
            "ratio_oferta_referencia_mediana",
            "unidad_precio_dominante",
            "precio_unitario_ref_mediana",
            "Rango unitario de referencia",
            "costo_objetivo_unitario",
            "nota_viabilidad_precio",
            "score",
            "Por qué destaca",
            "enlace_minsa",
        ]
    )
    exported = top[columns].rename(
        columns={
            "ficha": "Código de Ficha",
            "descripcion_oficial": "Descripción Oficial",
            "validacion_deteccion": "Validación de Detección",
            "actos_descartados_ambiguos": "Actos Descartados por Ambigüedad",
            "actos": "Cantidad de Actos",
            "actos_ficha_unica": "Actos de Ficha Única",
            "monto_total": "Monto Total Acumulado (USD)",
            "monto_ficha_unica": "Monto Ficha Única (USD)",
            "promedio_por_acto": "Promedio por Acto (USD)",
            "competencia": "Nivel de Competencia",
            "participantes_promedio": "Proponentes Promedio",
            "participantes_mediana": "Proponentes Mediana",
            "estatus_adjudicacion": "Estatus de Adjudicación",
            "requisitos": "Requisitos Exigidos",
            "clasificacion_oficial": "Clasificación Oficial",
            "clase_riesgo": "Clase de Riesgo",
            "presion_precio": "Presión Competitiva de Precio",
            "viabilidad_margen": "Viabilidad Preliminar de Margen",
            "propuestas_observadas": "Propuestas Comparables",
            "ratio_oferta_referencia_mediana": "Oferta/Referencia Típica",
            "unidad_precio_dominante": "Unidad de Precio Dominante",
            "precio_unitario_ref_mediana": "Precio Unitario Ref. Mediana (USD)",
            "costo_objetivo_unitario": "Costo Puesto Objetivo/Unidad (USD)",
            "nota_viabilidad_precio": "Lectura Comercial de Precio",
            "score": "Score Estratégico",
            "enlace_minsa": "Enlace MINSA",
        }
    )
    return exported


def _validate_rankings(
    exports: dict[str, pd.DataFrame],
    new_cutoff: pd.Timestamp,
) -> list[str]:
    checks: list[str] = []
    for sheet_name, frame in exports.items():
        assert len(frame) == TOP_N, f"{sheet_name}: se esperaban {TOP_N} filas y hay {len(frame)}"
        assert frame["Código de Ficha"].nunique() == TOP_N, f"{sheet_name}: fichas duplicadas"
        assert not set(frame["Código de Ficha"]) & EXCLUDED_FICHAS, f"{sheet_name}: exclusión fallida"
        assert frame["Requisitos Exigidos"].str.contains(r"RS: No", regex=True).all(), (
            f"{sheet_name}: se encontró una ficha que exige RS o no tiene el requisito confirmado"
        )
        assert not frame["Descripción Oficial"].map(normalize_text).map(
            lambda text: all(marker in text for marker in PEROXIDE_COMPLETE_MARKERS)
        ).any(), f"{sheet_name}: se encontró ciclo completo de peróxido"
        assert frame["Score Estratégico"].is_monotonic_decreasing, f"{sheet_name}: score desordenado"
        assert (frame["Monto Total Acumulado (USD)"] >= 0).all(), f"{sheet_name}: monto negativo"
        assert "Viabilidad Preliminar de Margen" in frame.columns
        assert frame["Validación de Detección"].astype(str).str.strip().ne("").all()
        checks.append(
            f"{sheet_name}: {TOP_N} fichas únicas, RS=No, exclusiones, evidencia, precio y orden correctos"
        )
    barrier = exports["3_Barrera_Cero"]
    assert barrier["Requisitos Exigidos"].str.contains(r"CT: No \| RS: No", regex=True).all()
    checks.append(f"3_Barrera_Cero: CT=No y RS=No confirmado en las {TOP_N} filas")
    new = exports["2_Nuevas_Potencial"]
    new_dates = pd.to_datetime(new["Fecha publicación/creación"], errors="coerce")
    assert new_dates.notna().all() and (new_dates >= new_cutoff).all()
    checks.append(f"2_Nuevas_Potencial: las {TOP_N} fichas son posteriores a {new_cutoff.date()}")
    deserted = exports["4_Actos_Desiertos"]
    assert deserted["Estatus de Adjudicación"].str.contains("desiertos", case=False).all()
    checks.append("4_Actos_Desiertos: universo construido exclusivamente con actos desiertos")
    ranked_codes = {
        normalize_ficha(value)
        for frame in exports.values()
        for value in frame["Código de Ficha"]
    }
    assert "107110" not in ranked_codes, "CILINDRO no superó la validación contextual"
    assert "107044" not in ranked_codes, "CUÑAS no superó la validación contextual"
    checks.append(
        "Control de falsos positivos: CILINDRO y CUÑAS quedaron fuera por falta de evidencia técnica suficiente"
    )
    return checks


def _style_excel(path: Path, built_at: str) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    top_fill = PatternFill("solid", fgColor="FFF2CC")
    top_font = Font(color="7F6000", bold=True)
    money_headers = {
        "Monto Total Acumulado (USD)",
        "Monto Ficha Única (USD)",
        "Promedio por Acto (USD)",
        "Precio Unitario Ref. Mediana (USD)",
        "Costo Puesto Objetivo/Unidad (USD)",
    }
    comments = {
        "Monto Total Acumulado (USD)": (
            "Suma del precio de referencia completo de cada acto donde aparece la ficha. "
            "Es una medida del tamaño del mercado relacionado y puede incluir otros renglones/fichas."
        ),
        "Monto Ficha Única (USD)": (
            "Suma del precio de referencia únicamente en actos que contienen una sola ficha técnica distinta. "
            "Es la medida monetaria más confiable para atribuir demanda a la ficha."
        ),
        "Score Estratégico": (
            "Percentil ponderado específico de cada categoría. No es una garantía de adjudicación ni utilidad."
        ),
        "Oferta/Referencia Típica": (
            "Mediana de la relación entre oferta observada y precio de referencia en actos de ficha única. "
            "Un valor bajo indica fuerte descuento competitivo frente a la referencia."
        ),
        "Costo Puesto Objetivo/Unidad (USD)": (
            "Estimación conservadora: precio unitario de referencia por la relación típica oferta/referencia, "
            "multiplicado por 75%. Es el costo máximo puesto en Panamá para aspirar a 25% de margen bruto, "
            "antes de gastos comerciales adicionales."
        ),
        "Validación de Detección": (
            "Control de calidad aplicado antes del ranking. Los nombres oficiales genéricos solo conservan "
            "actos que muestran el código de ficha o contexto técnico secundario coherente."
        ),
        "Actos Descartados por Ambigüedad": (
            "Actos que el detector original relacionó por un nombre genérico, pero que no superaron la "
            "validación de código o contexto técnico. No intervienen en montos ni puntajes."
        ),
    }
    for index, sheet_name in enumerate(SHEET_NAMES, start=1):
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.value in comments:
                cell.comment = Comment(comments[cell.value], "Codex")
        worksheet.row_dimensions[1].height = 40
        if worksheet.max_row >= 2:
            table = Table(displayName=f"TopOportunidades{index}", ref=worksheet.dimensions)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)
        headers = {cell.value: cell.column for cell in worksheet[1]}
        for row in range(2, worksheet.max_row + 1):
            rank = worksheet.cell(row, headers["Ranking"]).value
            if isinstance(rank, int) and rank <= 5:
                for cell in worksheet[row]:
                    cell.fill = top_fill
                worksheet.cell(row, headers["Prioridad"]).font = top_font
            worksheet.row_dimensions[row].height = 64
            for cell in worksheet[row]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for header in money_headers:
            if header in headers:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row, headers[header]).number_format = '$#,##0.00'
        for header in ("Cantidad de Actos", "Actos de Ficha Única"):
            if header in headers:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row, headers[header]).number_format = '#,##0'
        if "Score Estratégico" in headers:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row, headers["Score Estratégico"]).number_format = '0.0'
        if "Oferta/Referencia Típica" in headers:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row, headers["Oferta/Referencia Típica"]).number_format = '0.0%'
        if "Enlace MINSA" in headers:
            column = headers["Enlace MINSA"]
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, column)
                url = clean_text(cell.value)
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.value = "Abrir ficha oficial"
                    cell.style = "Hyperlink"
        for column_cells in worksheet.columns:
            header = clean_text(column_cells[0].value)
            values = [
                clean_text(cell.value)
                for cell in column_cells[: min(TOP_N + 1, len(column_cells))]
            ]
            width = min(max(max((len(value) for value in values), default=0) + 2, 11), 58)
            if header in {
                "Descripción Oficial",
                "Por qué destaca",
                "Estatus de Adjudicación",
                "Lectura Comercial de Precio",
                "Validación de Detección",
            }:
                width = 52
            elif header in {
                "Nivel de Competencia",
                "Requisitos Exigidos",
                "Clasificación Oficial",
                "Viabilidad Preliminar de Margen",
                "Rango unitario de referencia",
            }:
                width = 30
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
        worksheet.oddFooter.center.text = f"Fuente analítica construida: {built_at} | Perfil de detección ≥ {DETECTION_THRESHOLD:.0f}"
        worksheet.oddFooter.right.text = "Página &P de &N"
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.sheet_view.zoomScale = 80
    workbook.properties.title = "Top de oportunidades médicas en Panamá Compra"
    workbook.properties.subject = "Minería estratégica de fichas técnicas médicas"
    workbook.properties.creator = "GEAPP / Codex"
    workbook.save(path)


def generate_report(paths: InputPaths, output_path: Path) -> tuple[dict[str, pd.DataFrame], list[str]]:
    facts, metadata_raw, build_metadata = _load_source(paths)
    metadata = _derive_medical_scope(metadata_raw)
    price_intelligence = _load_price_intelligence(
        paths,
        set(metadata["ficha"].dropna().map(normalize_ficha)),
    )
    scoped_facts = _prepare_facts(facts, metadata)
    aggregate = _enrich_price_viability(_aggregate(scoped_facts), price_intelligence)
    cutoff = pd.Timestamp(date.today() - timedelta(days=730))

    rankings = {
        "1_Historicas": _score_historical(aggregate),
        "2_Nuevas_Potencial": _score_new(scoped_facts, cutoff, price_intelligence),
        "3_Barrera_Cero": _score_barrier_zero(aggregate),
        "4_Actos_Desiertos": _score_deserted(scoped_facts, price_intelligence),
    }
    categories = {
        "1_Historicas": "historical",
        "2_Nuevas_Potencial": "new",
        "3_Barrera_Cero": "barrier",
        "4_Actos_Desiertos": "deserted",
    }
    exports = {
        sheet: _to_export(rankings[sheet], categories[sheet]) for sheet in SHEET_NAMES
    }
    checks = _validate_rankings(exports, cutoff)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name in SHEET_NAMES:
            exports[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
    _style_excel(output_path, build_metadata.get("built_at_utc", "sin dato"))

    workbook = load_workbook(output_path, read_only=False, data_only=False)
    assert workbook.sheetnames == list(SHEET_NAMES), "El libro debe contener exactamente cuatro hojas"
    for sheet_name in SHEET_NAMES:
        worksheet = workbook[sheet_name]
        assert worksheet.max_row == TOP_N + 1, f"{sheet_name}: filas Excel inesperadas"
        assert len(worksheet.tables) == 1, f"{sheet_name}: falta la tabla con filtros"
        assert worksheet.freeze_panes == "A2", f"{sheet_name}: panel no congelado"
    workbook.close()
    checks.append(
        f"Excel reabierto: 4 hojas, {TOP_N} filas por hoja, filtros y paneles congelados verificados"
    )
    checks.append(
        "Auditoría de ambigüedad: "
        f"{int(scoped_facts.attrs.get('quality_dropped_relations', 0)):,} relaciones y "
        f"{int(scoped_facts.attrs.get('quality_dropped_acts', 0)):,} actos potencialmente ruidosos excluidos"
    )
    checks.append(
        f"Cobertura fuente: {len(facts):,} relaciones moderadas; {len(scoped_facts):,} médicas elegibles; "
        f"corte analítico {build_metadata.get('built_at_utc', 'sin dato')}"
    )
    checks.append(
        f"Inteligencia de precio: {len(price_intelligence):,} fichas con ofertas y/o precios unitarios observados"
    )
    return exports, checks


def parse_args() -> argparse.Namespace:
    home = Path.home()
    default_output = (
        home
        / "Downloads"
        / f"Top_Oportunidades_Medicas_{date.today().isoformat()}_Sin_RS_Top50.xlsx"
    )
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--analytics-db",
        type=Path,
        default=home / "scrapers_repo" / "data" / "db" / "inteligencia_proveedores.db",
    )
    parser.add_argument(
        "--ctni-db",
        type=Path,
        default=home / "scrapers_repo" / "data" / "ctni" / "ctni_monitor.db",
    )
    parser.add_argument(
        "--operational-db",
        type=Path,
        default=home / "scrapers_repo" / "data" / "db" / "panamacompra.db",
    )
    parser.add_argument("--output", type=Path, default=default_output)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    paths = InputPaths(
        args.analytics_db.resolve(),
        args.ctni_db.resolve(),
        args.operational_db.resolve(),
    )
    for path in (paths.analytics_db, paths.ctni_db, paths.operational_db):
        if not path.exists():
            raise FileNotFoundError(f"No existe la fuente requerida: {path}")
    exports, checks = generate_report(paths, args.output.resolve())
    print(f"Excel generado: {args.output.resolve()}")
    for check in checks:
        print(f"[OK] {check}")
    for sheet_name, frame in exports.items():
        preview = ", ".join(
            f"{row['Ranking']}. {row['Código de Ficha']} ({row['Score Estratégico']:.1f})"
            for _, row in frame.head(5).iterrows()
        )
        print(f"[TOP5] {sheet_name}: {preview}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
