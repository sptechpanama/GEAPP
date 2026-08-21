from __future__ import annotations

"""Convierte el Top de oportunidades medicas en un informe PDF ejecutivo.

El Excel sigue siendo la fuente tabular completa. El PDF prioriza legibilidad
en una pantalla de iPad: formato A4 horizontal, Top 5 explicado y las cincuenta
posiciones de cada categoria en tablas de diez filas.
"""

import argparse
import re
import unicodedata
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import fitz
from openpyxl import load_workbook


PAGE_W = 841.89
PAGE_H = 595.28
MARGIN = 28.0
TOP_N = 50
TOP_HIGHLIGHT = 5
ROWS_PER_RANKING_PAGE = 10

NAVY = (15 / 255, 39 / 255, 71 / 255)
NAVY_2 = (24 / 255, 54 / 255, 93 / 255)
TEAL = (18 / 255, 145 / 255, 140 / 255)
TEAL_LIGHT = (225 / 255, 247 / 255, 244 / 255)
GOLD = (226 / 255, 164 / 255, 52 / 255)
GOLD_LIGHT = (1.0, 247 / 255, 218 / 255)
INK = (28 / 255, 40 / 255, 55 / 255)
MUTED = (92 / 255, 108 / 255, 126 / 255)
LINE = (209 / 255, 219 / 255, 229 / 255)
WHITE = (1.0, 1.0, 1.0)
RED = (177 / 255, 52 / 255, 52 / 255)


CATEGORY_INFO = {
    "1_Historicas": {
        "title": "Fichas consolidadas / historicas",
        "short": "Historicas",
        "purpose": (
            "Demanda repetida, recurrencia y volumen monetario. Se premia la historia util, "
            "pero se descuenta la presion de precio para no confundir mercado grande con margen sano."
        ),
    },
    "2_Nuevas_Potencial": {
        "title": "Fichas nuevas de alto potencial",
        "short": "Nuevas",
        "purpose": (
            "Fichas creadas o trabajadas en los ultimos dos anos que ya muestran traccion, "
            "monto y competencia abordable. Las posiciones inferiores pueden ser senales "
            "iniciales y requieren validacion comercial antes de invertir."
        ),
    },
    "3_Barrera_Cero": {
        "title": "Rapida entrada / barrera cero",
        "short": "Barrera cero",
        "purpose": (
            "Solo fichas con Registro Sanitario = No y Criterio Tecnico = No confirmados. "
            "Favorece una entrada comercial mas corta, sin asumir que la venta esta garantizada."
        ),
    },
    "4_Actos_Desiertos": {
        "title": "Oportunidades en actos desiertos",
        "short": "Desiertos",
        "purpose": (
            "Niches donde la demanda no fue satisfecha. Se priorizan recurrencia y monto, "
            "verificando que el desierto no se deba a un precio imposible o requisito inviable; "
            "las senales de un solo acto se presentan como exploratorias."
        ),
    },
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return "" if text.lower() in {"none", "nan", "null", "<na>"} else text


def norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean(value).lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", text)


def row_value(row: dict[str, Any], *candidates: str, default: Any = "") -> Any:
    normalized = {norm(key): value for key, value in row.items()}
    for candidate in candidates:
        key = norm(candidate)
        if key in normalized and normalized[key] not in (None, ""):
            return normalized[key]
    return default


def as_float(value: Any) -> float:
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0


def money(value: Any, compact: bool = False) -> str:
    number = as_float(value)
    if compact:
        absolute = abs(number)
        if absolute >= 1_000_000:
            return f"${number / 1_000_000:,.2f}M"
        if absolute >= 1_000:
            return f"${number / 1_000:,.1f}K"
    return f"${number:,.2f}"


def integer(value: Any) -> str:
    return f"{int(round(as_float(value))):,}"


def percentage(value: Any) -> str:
    number = as_float(value)
    return "Sin muestra" if number <= 0 else f"{number:.1%}"


def load_excel(path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    for worksheet in workbook.worksheets:
        headers = [clean(cell.value) for cell in next(worksheet.iter_rows())]
        result[worksheet.title] = [
            dict(zip(headers, [cell.value for cell in cells]))
            for cells in worksheet.iter_rows(min_row=2)
            if any(cell.value not in (None, "") for cell in cells)
        ]
    workbook.close()
    missing = set(CATEGORY_INFO) - set(result)
    if missing:
        raise ValueError(f"Faltan hojas requeridas: {sorted(missing)}")
    for sheet_name in CATEGORY_INFO:
        if len(result[sheet_name]) != TOP_N:
            raise ValueError(
                f"{sheet_name} debe contener {TOP_N} registros, contiene {len(result[sheet_name])}"
            )
    return result


def _text_width(text: str, size: float, font: str = "helv") -> float:
    return fitz.get_text_length(text, fontname=font, fontsize=size)


def wrap(text: Any, width: float, size: float, max_lines: int | None = None) -> list[str]:
    words = clean(text).split()
    if not words:
        return [""]
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if not current or _text_width(candidate, size) <= width:
            current = candidate
            continue
        lines.append(current)
        current = word
        if max_lines and len(lines) >= max_lines:
            break
    if current and (not max_lines or len(lines) < max_lines):
        lines.append(current)
    if max_lines and len(lines) == max_lines:
        full = " ".join(words)
        shown = " ".join(lines)
        if shown != full:
            last = lines[-1]
            while last and _text_width(last + "...", size) > width:
                last = last[:-1].rstrip()
            lines[-1] = last + "..."
    return lines


def draw_wrapped(
    page: fitz.Page,
    rect: fitz.Rect,
    text: Any,
    *,
    size: float = 9.0,
    color: tuple[float, float, float] = INK,
    bold: bool = False,
    max_lines: int | None = None,
    line_height: float | None = None,
) -> float:
    font = "hebo" if bold else "helv"
    height = line_height or size * 1.28
    y = rect.y0 + size
    for line in wrap(text, rect.width, size, max_lines=max_lines):
        page.insert_text((rect.x0, y), line, fontsize=size, fontname=font, color=color)
        y += height
        if y > rect.y1 + 0.1:
            break
    return y


def add_header(page: fitz.Page, title: str, section: str = "") -> None:
    page.draw_rect(fitz.Rect(0, 0, PAGE_W, 42), fill=NAVY, color=NAVY)
    page.insert_text((MARGIN, 27), title, fontname="hebo", fontsize=17, color=WHITE)
    if section:
        width = _text_width(section, 8.5, "helv")
        page.insert_text((PAGE_W - MARGIN - width, 25), section, fontsize=8.5, color=(0.78, 0.87, 0.94))


def add_footer(page: fitz.Page, page_number: int, source_label: str) -> None:
    y = PAGE_H - 20
    page.draw_line((MARGIN, y - 6), (PAGE_W - MARGIN, y - 6), color=LINE, width=0.6)
    page.insert_text((MARGIN, y + 5), source_label, fontsize=6.7, color=MUTED)
    marker = f"Pagina {page_number}"
    page.insert_text(
        (PAGE_W - MARGIN - _text_width(marker, 6.7), y + 5),
        marker,
        fontsize=6.7,
        color=MUTED,
    )


def new_page(doc: fitz.Document, title: str, section: str, source_label: str) -> fitz.Page:
    page = doc.new_page(width=PAGE_W, height=PAGE_H)
    add_header(page, title, section)
    add_footer(page, len(doc), source_label)
    return page


def stat_card(page: fitz.Page, rect: fitz.Rect, label: str, value: str, note: str = "") -> None:
    page.draw_rect(rect, fill=(0.965, 0.98, 0.99), color=LINE, radius=0.08)
    page.insert_text((rect.x0 + 12, rect.y0 + 18), label.upper(), fontsize=6.8, color=MUTED, fontname="hebo")
    page.insert_text((rect.x0 + 12, rect.y0 + 43), value, fontsize=16, color=NAVY, fontname="hebo")
    if note:
        draw_wrapped(
            page,
            fitz.Rect(rect.x0 + 12, rect.y0 + 51, rect.x1 - 10, rect.y1 - 7),
            note,
            size=6.5,
            color=MUTED,
            max_lines=2,
        )


def top_five_codes(data: dict[str, list[dict[str, Any]]]) -> list[tuple[str, str, str]]:
    result: list[tuple[str, str, str]] = []
    for sheet, rows in data.items():
        for row in rows[:5]:
            result.append(
                (
                    clean(row_value(row, "Codigo de Ficha")),
                    clean(row_value(row, "Descripcion Oficial")),
                    CATEGORY_INFO[sheet]["short"],
                )
            )
    return result


def render_cover(doc: fitz.Document, source_label: str) -> None:
    page = doc.new_page(width=PAGE_W, height=PAGE_H)
    page.draw_rect(page.rect, fill=NAVY, color=NAVY)
    page.draw_rect(fitz.Rect(0, 0, 13, PAGE_H), fill=TEAL, color=TEAL)
    page.draw_rect(fitz.Rect(13, 0, 21, PAGE_H), fill=GOLD, color=GOLD)
    page.insert_text((58, 92), "INFORME ESTRATEGICO", fontsize=12, fontname="hebo", color=GOLD)
    draw_wrapped(
        page,
        fitz.Rect(58, 120, 750, 235),
        "Top de oportunidades medicas sin Registro Sanitario",
        size=31,
        color=WHITE,
        bold=True,
        max_lines=3,
        line_height=37,
    )
    draw_wrapped(
        page,
        fitz.Rect(60, 250, 730, 325),
        "Cuatro Top 50 recalculados desde la fuente analitica completa, con control de requisitos, "
        "calidad de deteccion, presion competitiva de precio y viabilidad preliminar de margen.",
        size=13,
        color=(0.82, 0.90, 0.96),
        max_lines=4,
        line_height=18,
    )
    page.draw_rect(fitz.Rect(58, 362, 782, 468), fill=NAVY_2, color=(0.2, 0.37, 0.55), radius=0.06)
    bullets = (
        "Registro Sanitario: solo fichas con NO confirmado",
        "Barrera cero: adicionalmente Criterio Tecnico = NO",
        "Excluye fichas ya revisadas y ciclo completo de peroxido",
        "Productos masivos: penalizados cuando el precio historico comprime el margen",
        "Nombres genericos: solo cuentan con codigo o contexto tecnico verificable",
    )
    y = 388
    for bullet in bullets:
        page.draw_circle((76, y - 3), 3.4, fill=TEAL, color=TEAL)
        page.insert_text((89, y), bullet, fontsize=10, color=WHITE)
        y += 17
    page.insert_text((58, 523), f"Fecha del informe: {date.today().isoformat()}", fontsize=9, color=(0.74, 0.84, 0.91))
    page.insert_text((58, 543), source_label, fontsize=7.2, color=(0.62, 0.74, 0.84))
    page.insert_text((PAGE_W - 108, 548), "GEAPP", fontsize=19, fontname="hebo", color=TEAL)


def render_methodology(doc: fitz.Document, source_label: str) -> None:
    page = new_page(doc, "Como se reconstruyeron los rankings", "Metodologia", source_label)
    x0, y0 = MARGIN, 64
    sections = [
        (
            "1. Universo elegible",
            "Solo insumos, dispositivos, equipos, instrumental e imagenologia medica. Se excluyen medicamentos, "
            "fichas con RS=Si, RS desconocido, los codigos ya revisados por el usuario y el ciclo completo de peroxido.",
        ),
        (
            "2. Demanda y dinero",
            "Cantidad de actos mide recurrencia. Monto de ficha unica es la señal monetaria mas atribuible; monto total "
            "mide el mercado relacionado y puede incluir otros renglones. Ambos se conservan para no ocultar contexto.",
        ),
        (
            "3. Precio y margen",
            "Las ofertas se comparan con su precio de referencia en actos de ficha unica. Los precios unitarios se agrupan "
            "por la unidad dominante para evitar mezclar caja, paquete, par y unidad. El costo puesto objetivo supone 25% "
            "de margen bruto y es una meta de negociacion, no una cotizacion garantizada.",
        ),
        (
            "4. Productos masivos",
            "Guantes, canulas, agujas, jeringas, gasas y similares reciben una penalizacion cuando el precio unitario o la "
            "relacion oferta/referencia evidencia guerra de precios. Por eso volumen alto no implica automaticamente Top 50.",
        ),
        (
            "5. Lectura correcta",
            "El ranking sirve para escoger donde invertir tiempo comercial. Antes de licitar se debe confirmar fabricante, "
            "MOQ, empaque, flete, aranceles, documentacion, plazo de pago y costo puesto real.",
        ),
    ]
    for index, (heading, body) in enumerate(sections):
        rect = fitz.Rect(x0 + (index % 2) * 393, y0 + (index // 2) * 151, x0 + (index % 2) * 393 + 372, y0 + (index // 2) * 151 + 128)
        page.draw_rect(rect, fill=WHITE, color=LINE, radius=0.06)
        page.draw_rect(fitz.Rect(rect.x0, rect.y0, rect.x0 + 6, rect.y1), fill=TEAL if index != 3 else GOLD, color=TEAL if index != 3 else GOLD)
        page.insert_text((rect.x0 + 18, rect.y0 + 25), heading, fontsize=11, fontname="hebo", color=NAVY)
        draw_wrapped(page, fitz.Rect(rect.x0 + 18, rect.y0 + 37, rect.x1 - 14, rect.y1 - 10), body, size=8.2, color=INK, max_lines=7)
    page.draw_rect(fitz.Rect(421, 366, 793, 494), fill=GOLD_LIGHT, color=GOLD, radius=0.06)
    page.insert_text((439, 391), "Regla conservadora de deteccion", fontsize=11, fontname="hebo", color=NAVY)
    draw_wrapped(
        page,
        fitz.Rect(439, 404, 775, 482),
        "Una ficha sin Registro Sanitario = NO confirmado queda fuera. Ademas, nombres genericos como CILINDRO o "
        "CUÑAS solo conservan actos respaldados por codigo oficial o contexto tecnico secundario; si la evidencia "
        "no alcanza el minimo comercial, la ficha no entra al ranking.",
        size=8.8,
        color=INK,
        max_lines=6,
    )


def render_executive_summary(
    doc: fitz.Document,
    data: dict[str, list[dict[str, Any]]],
    source_label: str,
) -> None:
    page = new_page(doc, "Resumen ejecutivo", "Top 5 por categoria", source_label)
    y = 64
    for index, (sheet, info) in enumerate(CATEGORY_INFO.items()):
        rows = data[sheet][:5]
        rect = fitz.Rect(MARGIN + (index % 2) * 393, y + (index // 2) * 230, MARGIN + (index % 2) * 393 + 372, y + (index // 2) * 230 + 207)
        page.draw_rect(rect, fill=(0.975, 0.985, 0.99), color=LINE, radius=0.06)
        page.draw_rect(fitz.Rect(rect.x0, rect.y0, rect.x1, rect.y0 + 34), fill=NAVY_2, color=NAVY_2, radius=0.06)
        page.insert_text((rect.x0 + 13, rect.y0 + 22), info["title"], fontsize=10.2, fontname="hebo", color=WHITE)
        yy = rect.y0 + 52
        for row in rows:
            rank = integer(row_value(row, "Ranking"))
            code = clean(row_value(row, "Codigo de Ficha"))
            desc = clean(row_value(row, "Descripcion Oficial"))
            score = as_float(row_value(row, "Score Estrategico"))
            page.insert_text((rect.x0 + 13, yy), f"{rank}. {code}", fontsize=8.2, fontname="hebo", color=TEAL)
            draw_wrapped(page, fitz.Rect(rect.x0 + 73, yy - 9, rect.x1 - 49, yy + 15), desc, size=7.5, color=INK, max_lines=1)
            page.insert_text((rect.x1 - 42, yy), f"{score:.1f}", fontsize=7.6, fontname="hebo", color=NAVY)
            yy += 29


def render_convergence(
    doc: fitz.Document,
    data: dict[str, list[dict[str, Any]]],
    source_label: str,
) -> None:
    page = new_page(doc, "Convergencia entre los cuatro rankings", "Senales repetidas", source_label)
    entries = top_five_codes(data)
    counts = Counter(code for code, _, _ in entries)
    repeated = sorted(counts.items(), key=lambda pair: (-pair[1], pair[0]))
    cards = []
    for code, count in repeated:
        desc = next(desc for c, desc, _ in entries if c == code)
        cats = sorted({cat for c, _, cat in entries if c == code})
        cards.append((code, desc, count, ", ".join(cats)))
    page.insert_text((MARGIN, 68), "Una ficha repetida en varios Top 5 combina mas de una señal favorable.", fontsize=10, color=MUTED)
    if not cards:
        cards = [("-", "No hubo coincidencias entre Top 5; revisar cada categoria por separado.", 0, "")]
    y = 92
    for index, (code, desc, count, cats) in enumerate(cards[:9]):
        rect = fitz.Rect(MARGIN + (index % 3) * 263, y + (index // 3) * 130, MARGIN + (index % 3) * 263 + 244, y + (index // 3) * 130 + 111)
        page.draw_rect(rect, fill=TEAL_LIGHT if count > 1 else WHITE, color=TEAL if count > 1 else LINE, radius=0.06)
        page.insert_text((rect.x0 + 13, rect.y0 + 23), code, fontsize=12, fontname="hebo", color=NAVY)
        page.insert_text((rect.x1 - 66, rect.y0 + 22), f"{count} Top 5", fontsize=7.5, fontname="hebo", color=TEAL)
        draw_wrapped(page, fitz.Rect(rect.x0 + 13, rect.y0 + 34, rect.x1 - 12, rect.y0 + 72), desc, size=7.8, color=INK, max_lines=2)
        draw_wrapped(page, fitz.Rect(rect.x0 + 13, rect.y0 + 76, rect.x1 - 12, rect.y1 - 8), cats, size=6.8, color=MUTED, max_lines=2)


def category_metrics(rows: list[dict[str, Any]]) -> dict[str, str]:
    return {
        "Actos": integer(sum(as_float(row_value(row, "Cantidad de Actos")) for row in rows)),
        "Monto total": money(sum(as_float(row_value(row, "Monto Total Acumulado USD")) for row in rows), compact=True),
        "Monto ficha unica": money(sum(as_float(row_value(row, "Monto Ficha Unica USD")) for row in rows), compact=True),
        "Score medio": f"{sum(as_float(row_value(row, 'Score Estrategico')) for row in rows) / len(rows):.1f}",
    }


def render_category_overview(
    doc: fitz.Document,
    sheet: str,
    rows: list[dict[str, Any]],
    source_label: str,
) -> None:
    info = CATEGORY_INFO[sheet]
    page = new_page(doc, info["title"], "Vista general", source_label)
    draw_wrapped(page, fitz.Rect(MARGIN, 58, PAGE_W - MARGIN, 91), info["purpose"], size=9.1, color=MUTED, max_lines=2)
    metrics = category_metrics(rows)
    x = MARGIN
    for label, value in metrics.items():
        stat_card(page, fitz.Rect(x, 101, x + 184, 178), label, value)
        x += 196
    page.insert_text((MARGIN, 211), "Top 5 y motivo estrategico", fontsize=13, fontname="hebo", color=NAVY)
    y = 231
    for row in rows[:5]:
        rank = integer(row_value(row, "Ranking"))
        code = clean(row_value(row, "Codigo de Ficha"))
        desc = clean(row_value(row, "Descripcion Oficial"))
        why = clean(row_value(row, "Por que destaca"))
        score = as_float(row_value(row, "Score Estrategico"))
        page.draw_rect(fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + 54), fill=WHITE, color=LINE, radius=0.06)
        page.draw_rect(fitz.Rect(MARGIN, y, MARGIN + 44, y + 54), fill=TEAL if int(rank) <= 3 else NAVY_2, color=TEAL if int(rank) <= 3 else NAVY_2, radius=0.06)
        page.insert_text((MARGIN + 15, y + 32), rank, fontsize=16, fontname="hebo", color=WHITE)
        page.insert_text((MARGIN + 56, y + 18), f"{code} | {desc[:76]}", fontsize=8.5, fontname="hebo", color=NAVY)
        draw_wrapped(page, fitz.Rect(MARGIN + 56, y + 25, PAGE_W - MARGIN - 75, y + 49), why, size=7.2, color=INK, max_lines=2)
        page.insert_text((PAGE_W - MARGIN - 48, y + 31), f"{score:.1f}", fontsize=10, fontname="hebo", color=TEAL)
        y += 61


def render_top_detail_pages(
    doc: fitz.Document,
    sheet: str,
    rows: list[dict[str, Any]],
    source_label: str,
) -> None:
    info = CATEGORY_INFO[sheet]
    chunks = (rows[:3], rows[3:5])
    for page_index, chunk in enumerate(chunks, start=1):
        page = new_page(doc, f"{info['title']} - Top 5", f"Detalle {page_index}/2", source_label)
        y = 59
        card_height = 160 if len(chunk) == 3 else 218
        for row in chunk:
            rank = integer(row_value(row, "Ranking"))
            code = clean(row_value(row, "Codigo de Ficha"))
            desc = clean(row_value(row, "Descripcion Oficial"))
            acts = integer(row_value(row, "Cantidad de Actos"))
            unique = integer(row_value(row, "Actos de Ficha Unica"))
            total = money(row_value(row, "Monto Total Acumulado USD"), compact=True)
            unique_amount = money(row_value(row, "Monto Ficha Unica USD"), compact=True)
            competition = clean(row_value(row, "Nivel de Competencia"))
            reqs = clean(row_value(row, "Requisitos Exigidos"))
            viability = clean(row_value(row, "Viabilidad Preliminar de Margen"))
            pressure = clean(row_value(row, "Presion Competitiva de Precio"))
            price_note = clean(row_value(row, "Lectura Comercial de Precio"))
            why = clean(row_value(row, "Por que destaca"))
            score = as_float(row_value(row, "Score Estrategico"))
            official_url = clean(row_value(row, "Enlace MINSA"))
            rect = fitz.Rect(MARGIN, y, PAGE_W - MARGIN, y + card_height - 9)
            page.draw_rect(rect, fill=WHITE, color=LINE, radius=0.06)
            page.draw_rect(fitz.Rect(rect.x0, rect.y0, rect.x0 + 66, rect.y1), fill=NAVY_2, color=NAVY_2, radius=0.06)
            page.insert_text((rect.x0 + 18, rect.y0 + 34), f"#{rank}", fontsize=18, fontname="hebo", color=WHITE)
            page.insert_text((rect.x0 + 18, rect.y0 + 61), code, fontsize=10, fontname="hebo", color=GOLD)
            page.insert_text((rect.x0 + 18, rect.y0 + 91), "SCORE", fontsize=6.4, fontname="hebo", color=(0.7, 0.82, 0.91))
            page.insert_text((rect.x0 + 18, rect.y0 + 113), f"{score:.1f}", fontsize=16, fontname="hebo", color=TEAL_LIGHT)
            left = rect.x0 + 80
            page.insert_text((left, rect.y0 + 21), desc, fontsize=9.2, fontname="hebo", color=NAVY)
            labels = [
                ("Actos", acts),
                ("Ficha unica", unique),
                ("Mercado relacionado", total),
                ("Monto ficha unica", unique_amount),
            ]
            xx = left
            for label, value in labels:
                page.insert_text((xx, rect.y0 + 44), label.upper(), fontsize=5.9, fontname="hebo", color=MUTED)
                page.insert_text((xx, rect.y0 + 60), value, fontsize=9.5, fontname="hebo", color=TEAL)
                xx += 132
            page.insert_text((left, rect.y0 + 82), f"Competencia: {competition}", fontsize=7.1, color=INK)
            page.insert_text((left + 330, rect.y0 + 82), f"Precio: {pressure} | {viability}", fontsize=7.1, color=INK)
            draw_wrapped(page, fitz.Rect(left, rect.y0 + 91, rect.x1 - 12, rect.y0 + 122), f"Por que destaca: {why}", size=6.9, color=INK, max_lines=2)
            draw_wrapped(page, fitz.Rect(left, rect.y0 + 120, rect.x1 - 12, rect.y0 + 150), f"Lectura de precio: {price_note}", size=6.6, color=MUTED, max_lines=2)
            if card_height > 180:
                draw_wrapped(page, fitz.Rect(left, rect.y0 + 157, rect.x1 - 12, rect.y0 + 186), f"Requisitos: {reqs}", size=7.0, color=INK, max_lines=2)
                if official_url.startswith("http"):
                    link_rect = fitz.Rect(left, rect.y0 + 191, left + 105, rect.y0 + 211)
                    page.draw_rect(link_rect, fill=TEAL_LIGHT, color=TEAL, radius=0.08)
                    page.insert_text((link_rect.x0 + 10, link_rect.y0 + 14), "Abrir ficha MINSA", fontsize=7.1, fontname="hebo", color=TEAL)
                    page.insert_link({"kind": fitz.LINK_URI, "from": link_rect, "uri": official_url})
            y += card_height


def _table_cell(
    page: fitz.Page,
    rect: fitz.Rect,
    text: Any,
    *,
    size: float,
    bold: bool = False,
    color: tuple[float, float, float] = INK,
    max_lines: int = 2,
    align_right: bool = False,
) -> None:
    value = clean(text)
    if align_right:
        width = _text_width(value, size, "hebo" if bold else "helv")
        page.insert_text((max(rect.x0 + 2, rect.x1 - width - 4), rect.y0 + size + 3), value, fontsize=size, fontname="hebo" if bold else "helv", color=color)
    else:
        draw_wrapped(page, fitz.Rect(rect.x0 + 4, rect.y0 + 2, rect.x1 - 3, rect.y1 - 2), value, size=size, color=color, bold=bold, max_lines=max_lines, line_height=size * 1.18)


def render_ranking_pages(
    doc: fitz.Document,
    sheet: str,
    rows: list[dict[str, Any]],
    source_label: str,
) -> None:
    info = CATEGORY_INFO[sheet]
    headers = ["#", "Ficha / descripcion", "Actos", "Unicos", "Monto total", "Ficha unica", "Competencia", "Requisitos", "Precio / margen", "Score"]
    widths = [28, 210, 44, 42, 65, 65, 78, 92, 108, 42]
    for start in range(0, len(rows), ROWS_PER_RANKING_PAGE):
        end = min(start + ROWS_PER_RANKING_PAGE, len(rows))
        page = new_page(
            doc,
            f"{info['title']} - ranking completo",
            f"Posiciones {start + 1}-{end}",
            source_label,
        )
        x = MARGIN
        y = 68
        for header, width in zip(headers, widths):
            rect = fitz.Rect(x, y, x + width, y + 34)
            page.draw_rect(rect, fill=NAVY_2, color=WHITE, width=0.4)
            _table_cell(page, rect, header, size=6.5, bold=True, color=WHITE, max_lines=2)
            x += width
        y += 34
        for idx, row in enumerate(rows[start:end]):
            values = [
                integer(row_value(row, "Ranking")),
                f"{clean(row_value(row, 'Codigo de Ficha'))} | {clean(row_value(row, 'Descripcion Oficial'))}",
                integer(row_value(row, "Cantidad de Actos")),
                integer(row_value(row, "Actos de Ficha Unica")),
                money(row_value(row, "Monto Total Acumulado USD"), compact=True),
                money(row_value(row, "Monto Ficha Unica USD"), compact=True),
                clean(row_value(row, "Nivel de Competencia")),
                clean(row_value(row, "Requisitos Exigidos")),
                f"{clean(row_value(row, 'Presion Competitiva de Precio'))} | {clean(row_value(row, 'Viabilidad Preliminar de Margen'))}",
                f"{as_float(row_value(row, 'Score Estrategico')):.1f}",
            ]
            row_h = 43
            x = MARGIN
            fill = GOLD_LIGHT if start + idx < 5 else ((0.965, 0.98, 0.99) if idx % 2 == 0 else WHITE)
            for col_index, (value, width) in enumerate(zip(values, widths)):
                rect = fitz.Rect(x, y, x + width, y + row_h)
                page.draw_rect(rect, fill=fill, color=LINE, width=0.4)
                _table_cell(
                    page,
                    rect,
                    value,
                    size=6.2 if col_index not in {1, 7, 8} else 5.8,
                    bold=col_index in {0, 1, 9},
                    color=NAVY if col_index in {0, 1, 9} else INK,
                    max_lines=3 if col_index in {1, 7, 8} else 2,
                    align_right=col_index in {2, 3, 4, 5, 9},
                )
                x += width
            y += row_h


def render_action_plan(doc: fitz.Document, source_label: str) -> None:
    page = new_page(doc, "Ruta de validacion comercial", "Siguientes pasos", source_label)
    steps = [
        ("1", "Confirmar ficha", "Leer CT, presentacion, unidad, clase y especificaciones obligatorias."),
        ("2", "Mapear precios", "Revisar renglones comparables, mediana, cuartil bajo y empaque dominante."),
        ("3", "Buscar fabricante", "Solicitar MOQ, ficha del producto, certificaciones, plazo y exclusividad local."),
        ("4", "Calcular costo puesto", "Producto + flete + seguro + nacionalizacion + financiamiento + contingencia."),
        ("5", "Exigir margen", "Comparar costo puesto con el costo objetivo; no entrar si depende de supuestos fragiles."),
        ("6", "Piloto", "Atacar primero 3-5 fichas, medir respuestas y escalar solo las que validen margen real."),
    ]
    y = 75
    for index, (number, title, body) in enumerate(steps):
        rect = fitz.Rect(MARGIN + (index % 2) * 393, y + (index // 2) * 145, MARGIN + (index % 2) * 393 + 372, y + (index // 2) * 145 + 120)
        page.draw_rect(rect, fill=WHITE, color=LINE, radius=0.06)
        page.draw_circle((rect.x0 + 34, rect.y0 + 38), 20, fill=TEAL, color=TEAL)
        page.insert_text((rect.x0 + 28, rect.y0 + 44), number, fontsize=15, fontname="hebo", color=WHITE)
        page.insert_text((rect.x0 + 66, rect.y0 + 29), title, fontsize=11, fontname="hebo", color=NAVY)
        draw_wrapped(page, fitz.Rect(rect.x0 + 66, rect.y0 + 40, rect.x1 - 14, rect.y1 - 12), body, size=8.1, color=INK, max_lines=5)
    page.draw_rect(fitz.Rect(MARGIN, 514, PAGE_W - MARGIN, 552), fill=GOLD_LIGHT, color=GOLD, radius=0.06)
    page.insert_text((MARGIN + 15, 537), "Decision recomendada: usar el Top 50 como embudo; usar cotizaciones reales para autorizar capital.", fontsize=9, fontname="hebo", color=NAVY)


def render_glossary(doc: fitz.Document, source_label: str) -> None:
    page = new_page(doc, "Glosario y limites", "Cierre", source_label)
    items = [
        ("Actos", "Cantidad de procesos distintos donde la ficha fue detectada con el perfil analitico moderado."),
        ("Ficha unica", "Actos donde solo se detecto una ficha tecnica distinta; es la base monetaria mas atribuible."),
        ("Monto total", "Precio de referencia completo de actos relacionados; puede contener otros renglones."),
        ("Oferta/referencia", "Relacion observada entre oferta y precio de referencia en actos comparables."),
        ("Costo puesto objetivo", "Techo estimado para buscar 25% de margen bruto; no incluye todos los riesgos comerciales."),
        ("Presion de precio", "Clasificacion basada en descuentos historicos y, para masivos, en el precio unitario."),
        ("Viabilidad preliminar", "Lectura cuantitativa previa a recibir una cotizacion vinculante del fabricante."),
        ("RS confirmado", "Solo se incluyeron fichas cuyo metadato oficial indica Registro Sanitario = No."),
    ]
    y = 70
    for index, (term, definition) in enumerate(items):
        rect = fitz.Rect(MARGIN + (index % 2) * 393, y + (index // 2) * 109, MARGIN + (index % 2) * 393 + 372, y + (index // 2) * 109 + 89)
        page.draw_rect(rect, fill=(0.975, 0.985, 0.99), color=LINE, radius=0.06)
        page.insert_text((rect.x0 + 13, rect.y0 + 23), term, fontsize=9.5, fontname="hebo", color=TEAL)
        draw_wrapped(page, fitz.Rect(rect.x0 + 13, rect.y0 + 31, rect.x1 - 12, rect.y1 - 8), definition, size=7.5, color=INK, max_lines=4)
    page.draw_rect(fitz.Rect(MARGIN, 512, PAGE_W - MARGIN, 552), fill=(0.98, 0.94, 0.94), color=RED, radius=0.06)
    page.insert_text((MARGIN + 14, 536), "Este informe prioriza investigacion comercial; no sustituye validacion regulatoria, tecnica, financiera ni contractual.", fontsize=8.7, fontname="hebo", color=RED)


def validate_pdf(path: Path, data: dict[str, list[dict[str, Any]]]) -> list[str]:
    doc = fitz.open(path)
    checks: list[str] = []
    expected_min_pages = 6 + len(CATEGORY_INFO) * (
        3 + ((TOP_N + ROWS_PER_RANKING_PAGE - 1) // ROWS_PER_RANKING_PAGE)
    )
    if doc.page_count < expected_min_pages:
        raise AssertionError(f"PDF demasiado corto: {doc.page_count} paginas")
    all_text = "\n".join(page.get_text("text") for page in doc)
    for page_index, page in enumerate(doc, start=1):
        if page.rect.width <= page.rect.height:
            raise AssertionError(f"Pagina {page_index} no esta en formato horizontal")
        if len(page.get_text("text").strip()) < 80:
            raise AssertionError(f"Pagina {page_index} parece vacia")
    expected_codes = {
        clean(row_value(row, "Codigo de Ficha"))
        for rows in data.values()
        for row in rows
    }
    missing = sorted(code for code in expected_codes if code and code not in all_text)
    if missing:
        raise AssertionError(f"Codigos ausentes en PDF: {missing}")
    checks.append(f"{doc.page_count} paginas A4 horizontales verificadas")
    checks.append(f"Los {len(expected_codes)} codigos unicos del Excel aparecen en el PDF")
    checks.append("Todas las paginas contienen texto seleccionable y no estan vacias")
    doc.close()
    return checks


def generate_pdf(excel_path: Path, output_path: Path) -> list[str]:
    data = load_excel(excel_path)
    source_label = f"Fuente: {excel_path.name} | Perfil >= 90 | RS=No confirmado"
    doc = fitz.open()
    render_cover(doc, source_label)
    render_methodology(doc, source_label)
    render_executive_summary(doc, data, source_label)
    render_convergence(doc, data, source_label)
    for sheet, rows in data.items():
        render_category_overview(doc, sheet, rows, source_label)
        render_top_detail_pages(doc, sheet, rows, source_label)
        render_ranking_pages(doc, sheet, rows, source_label)
    render_action_plan(doc, source_label)
    render_glossary(doc, source_label)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc.set_metadata(
        {
            "title": "Top de oportunidades medicas sin Registro Sanitario",
            "author": "GEAPP / Codex",
            "subject": "Cuatro Top 50 recalculados y auditados",
            "keywords": "Panama Compra, oportunidades medicas, fichas tecnicas, margen",
        }
    )
    doc.save(output_path, garbage=4, deflate=True, clean=True)
    doc.close()
    return validate_pdf(output_path, data)


def parse_args() -> argparse.Namespace:
    home = Path.home()
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--excel",
        type=Path,
        default=(
            home
            / "Downloads"
            / f"Top_Oportunidades_Medicas_{date.today().isoformat()}_Sin_RS_Top50.xlsx"
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=(
            home
            / "Downloads"
            / f"Informe_Oportunidades_Medicas_{date.today().isoformat()}_Sin_RS_Top50.pdf"
        ),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    excel = args.excel.resolve()
    output = args.output.resolve()
    if not excel.exists():
        raise FileNotFoundError(f"No existe el Excel fuente: {excel}")
    checks = generate_pdf(excel, output)
    print(f"PDF generado: {output}")
    for check in checks:
        print(f"[OK] {check}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
